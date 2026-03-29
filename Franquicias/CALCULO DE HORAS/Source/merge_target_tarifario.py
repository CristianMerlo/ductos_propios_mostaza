import openpyxl
import os
import json
from difflib import get_close_matches

# ─────────────────────────────────────────────────────────
# MERGE TARIFARIO → TARGET
# Agrega columnas de zona, km y tarifa al target.xlsx
# basándose en el Cód. de Sucursal de cada fila.
# ─────────────────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# El script está en Source/, subimos un nivel para encontrar Tarifario/
BASE_DIR = os.path.dirname(SCRIPT_DIR)

TARIFARIO_PATH = os.path.join(BASE_DIR, "Tarifario", "Clasificacion de Locales.xlsx")
TARGET_PATH    = os.path.join(SCRIPT_DIR, "target.xlsx")
SCHEMA_PATH    = os.path.join(SCRIPT_DIR, "SCHEMA.json")

def load_schema():
    if os.path.exists(SCHEMA_PATH):
        with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    print("⚠️  No se encontró SCHEMA.json en Source/.")
    return {}

# ─────────────────────────────────────────────────────────
# Mapping manual: Cód. de Sucursal → nombre en tarifario
# ─────────────────────────────────────────────────────────
MAPPING_MANUAL = {
    "F3DF":   "SAN MARTIN AUTO",
    "FAVE":   "AVELLANEDA",
    "FCABI":  "CABILDO",
    "FCFQU":  "CARRE QUILMES",
    "FCSJ3":  "SAN JUSTO",
    "FEZEA":  "EZEIZA",
    "FFLOR":  "FLORES",
    "FGONN":  "GONNET",
    "FGRBG":  "GRAND BOURG",
    "FGUE":   "GUEMES",
    "FLF2":   "LAFERRERE 2",
    "FLINR":  "LINIERS",
    "FLOA":   "LOMAS AUTO",
    "FLPC4":  "LA PLATA 4",
    "FLPCA":  "LA PLATA 2",
    "FLUJA":  "LUJAN",
    "FMBER":  "BERAZATEGUI",
    "FMBLV":  "BOLIVAR",
    "FMBOE":  "BOEDO",
    "FMCAN":  "CANNING",
    "FMCNT":  "CONSTITUCION",
    "FMCPE":  "PLAZA ESPAÑA",
    "FMCYM":  "CABILDO 2",
    "FMDQ2":  "MAR DEL PLATA",
    "FMDQ4":  "MAR DEL PLATA",
    "FMDQG":  "MAR DEL PLATA",
    "FMERL":  "MERLO",
    "FMGRA":  "MONTEGRANDE",
    "FMGUA":  "GUALEGUAYCHU",
    "FMJCP":  "JOSE C PAZ",
    "FMJUN":  "JUNIN",
    "FMLAC":  "LANUS 2",
    "FMLCA":  "LOMAS",
    "FMMSM":  "MENDOZA",
    "FMMRE":  "MORENO",
    "FMONC":  "ONCE",
    "FMPCH":  "PARQUE CHAC",
    "FMP3":   "LA PERLA",
    "FMPRI":  "PRIMERA JUNTA",
    "FMPUM":  "ROSARIO SUR",
    "FMPYA":  "POMPEYA",
    "FMQCA":  "QUILMES P.",
    "FMRAM":  "RAMOS",
    "FMROS":  "PELLEGRINI",
    "FMSFV":  "SANTA FE",
    "FMSIS":  "SAN ISIDRO",
    "FMSMA":  "SAN MIGUEL A.",
    "FMSTO":  "SAN TELMO",
    "FMSVP":  "SAN MARTIN PEATONAL",
    "FMURQ":  "URQUIZA",
    "FMVCP":  "CARLOS PAZ",
    "FMWVP":  "WILDE",
    "FMZAR":  "ZARATE",
    "FORO":   "ROSARIO SUR",
    "FPMPI":  "PALMAS PILAR",
    "FPROS":  "PORTAL ROSARIO",
    "FROG":   "ROTONDA",
    "FRSM":   "ROSARIO SUR",
    "FSAF":   "SAN FERNANDO",
    "FSAO":   "SALTA ORAN",
    "FSANT":  "SANTIAGO",
    "FSLIB":  "SALTA LIBERTAD",
    "FSM3":   "SAN MIGUEL 3",
    "FTAND":  "TANDIL",
    "FUSHA":  "USHUAIA",
    "FVDP":   "VILLA DEL PARQUE",
}


def cargar_tarifario():
    """Carga el tarifario y devuelve dict: nombre_local → datos"""
    if not os.path.exists(TARIFARIO_PATH):
        print(f"❌ Error: No se encontró el tarifario en {TARIFARIO_PATH}")
        return {}
        
    wb = openpyxl.load_workbook(TARIFARIO_PATH)
    ws = wb.active # Asumimos la hoja activa
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    
    try:
        col_km  = next(i for i, h in enumerate(headers) if "km" in h.lower() or "kilom" in h.lower())
        col_rad = next(i for i, h in enumerate(headers) if "radio" in h.lower() or "zona" in h.lower())
        col_vh  = next(i for i, h in enumerate(headers) if "valor hora" in h.lower() or "tarifa" in h.lower())
        col_mnh = next(i for i, h in enumerate(headers) if "minimo" in h.lower() or "hs min" in h.lower())
    except StopIteration:
        print("❌ Error: Cabeceras del tarifario no reconocidas.")
        return {}

    datos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        local = str(row[0]).strip().upper() if row[0] else None
        if local:
            datos[local] = {
                "km":        row[col_km],
                "radio":     row[col_rad],
                "valor_hora": row[col_vh],
                "min_horas": row[col_mnh],
            }
    return datos


def zona_desde_km(km):
    if km is None or not isinstance(km, (int, float)):
        return "Sin dato / Interior"
    if km <= 30:
        return "Zona 1 — CABA / GBA Cercano (≤30 km)"
    elif km <= 60:
        return "Zona 2 — GBA Medio (30-60 km)"
    else:
        return "Zona 3+ — GBA Lejano / Interior (>60 km)"


def resolver_datos_sucursal(cod, nombre_suc, tarifario):
    tar_keys = list(tarifario.keys())
    local_key = MAPPING_MANUAL.get(cod)
    if local_key and local_key in tarifario:
        return tarifario[local_key]

    suc_upper = str(nombre_suc).strip().upper() if nombre_suc else ""
    close = get_close_matches(suc_upper, tar_keys, n=1, cutoff=0.4)
    if close:
        return tarifario[close[0]]

    return {"km": None, "radio": "Sin dato", "valor_hora": 35, "min_horas": 4}


def merge_tarifario_a_target():
    print(f"─" * 60)
    print(f"  INTEGRADOR DE TARIFARIO (Esquema Dinámico)")
    print(f"─" * 60)
    
    schema = load_schema()
    col_name_cod = schema.get('cod_sucursal', 'Cód. de Sucursal')
    col_name_suc = schema.get('sucursal', 'Sucursal')
    
    tarifario = cargar_tarifario()
    if not tarifario: return
    print(f"✅ Tarifario cargado: {len(tarifario)} locales")

    if not os.path.exists(TARGET_PATH):
        print(f"❌ No se encontró {TARGET_PATH}")
        return
        
    wb = openpyxl.load_workbook(TARGET_PATH)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    
    if col_name_cod not in headers or col_name_suc not in headers:
        print(f"❌ Error: Columnas '{col_name_cod}' o '{col_name_suc}' no encontradas.")
        return

    col_cod_idx = headers.index(col_name_cod) + 1
    col_suc_idx = headers.index(col_name_suc) + 1

    nuevas_cols = {
        "zona": schema.get('zona', 'Zona'),
        "km": schema.get('km', 'Km'),
        "valor_hora": schema.get('valor_hora', 'Valor Hora (USD)'),
        "hs_minimas": schema.get('hs_minimas', 'Hs Mínimas')
    }
    
    col_indices = {}
    for role, col_name in nuevas_cols.items():
        if col_name in headers:
            col_indices[role] = headers.index(col_name) + 1
        else:
            ws.cell(row=1, column=ws.max_column + 1, value=col_name)
            col_indices[role] = ws.max_column

    procesadas = 0
    for row_idx in range(2, ws.max_row + 1):
        cod = ws.cell(row=row_idx, column=col_cod_idx).value
        suc = ws.cell(row=row_idx, column=col_suc_idx).value
        if not cod: continue

        datos = resolver_datos_sucursal(str(cod), suc, tarifario)
        km, vh, mnh = datos["km"], datos["valor_hora"], datos["min_horas"]
        zona = zona_desde_km(km)

        ws.cell(row=row_idx, column=col_indices["zona"]).value = zona
        ws.cell(row=row_idx, column=col_indices["km"]).value = km
        ws.cell(row=row_idx, column=col_indices["valor_hora"]).value = vh
        ws.cell(row=row_idx, column=col_indices["hs_minimas"]).value = mnh
        procesadas += 1

    wb.save(TARGET_PATH)
    print(f"\n✅ target.xlsx actualizado: {procesadas} filas procesadas")


if __name__ == "__main__":
    merge_tarifario_a_target()
