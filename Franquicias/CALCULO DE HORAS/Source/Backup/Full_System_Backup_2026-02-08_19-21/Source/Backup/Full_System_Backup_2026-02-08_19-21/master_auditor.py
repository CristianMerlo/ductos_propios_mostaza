import pandas as pd
import math
import shutil
import re
import os
import glob
import json
from datetime import datetime
from pypdf import PdfReader

# --- CONFIGURACIÓN Y PARÁMETROS ---
SOURCE_DIR = "Source"
EXCEL_PATH = os.path.join(SOURCE_DIR, "Target.xlsx")
CHAT_PATTERN = os.path.join(SOURCE_DIR, "*chat*.txt")
PDF_DIR = os.path.join(SOURCE_DIR, "Informes")
BASE_TECNICOS_PATH = os.path.join(SOURCE_DIR, "Base Tecnicos.xlsx")
TARIFAS_PATH = os.path.join(SOURCE_DIR, "tarifas.txt")
BACKUP_DIR = os.path.join(SOURCE_DIR, "Backup")
MONTH_TO_EVAL = "enero" # Desde PARAMETROS.txt

# --- LÓGICA 1: AUDITORÍA TÉCNICA (REGLAS) ---
def get_technical_hours(row):
    text = f"{row['Título'] if 'Título' in row else ''} {row['Incidencia']} {row['Descripción']} {row['Última Respuesta']} {row['Mensaje de Resolución']}".lower()
    
    # Reglas Cimbali
    if any(err in text for err in ['54', '55']) and any(word in text for word in ['er', 'error']):
        return 5
    if '58' in text and any(word in text for word in ['er', 'error']):
        return 6
    if '59' in text:
        return 1
        
    # Equipos No-Café
    if any(w in text for w in ['freidora', 'locatelli']):
        if any(w in text for w in ['picos', 'quemadores', 'explos']): return 3
    if any(w in text for w in ['horno', 'unox']):
        if any(w in text for w in ['contactora', 'cable']): return 5
    if any(w in text for w in ['broiler', 'nieco']):
        return 2
        
    # Regla de Sabotaje / Trivial
    if any(w in text for w in ['destrabe', 'madera']):
        return 3
    if 'se cierra' in text and len(text.strip()) < 50:
        return 2

    # Tiempo de servicio explícito (Diferenciando min/hs y aplicando redondeo)
    exact_match = re.search(r'tiempo de servicio[:\s]+(\d+)\s*([a-z]+)?', text)
    if exact_match:
        val = int(exact_match.group(1))
        unit = exact_match.group(2) if exact_match.group(2) else 'hs'
        if unit.startswith('m'): 
            return math.ceil(val / 60)
        else: 
            return val

    if len(text.strip()) > 10: return 2
    return 0

# --- LÓGICA 2: AUDITORÍA DE CHAT ---
def parse_chats():
    reports = {} # {ticket_id: hours}
    chat_files = glob.glob(CHAT_PATTERN)
    msg_start = re.compile(r'^\[(\d+/\d+/\d+),.*\] (.*): (.*)')
    
    for chat_path in chat_files:
        current_report_text = ""
        current_ticket = None
        
        with open(chat_path, 'r', encoding='utf-8') as f:
            for line in f:
                match = msg_start.match(line)
                if match:
                    # Al encontrar un nuevo mensaje, procesamos el anterior si tenía ticket
                    if current_ticket and current_report_text:
                        hours = audit_text_minimal(current_report_text)
                        reports[current_ticket] = max(reports.get(current_ticket, 0), hours)
                    
                    _, _, content = match.groups()
                    current_report_text = content
                    # Intentar extraer ticket
                    code_match = re.search(r'tiket#\s*(\d+)', content.lower())
                    current_ticket = code_match.group(1) if code_match else None
                else:
                    current_report_text += " " + line.strip()
                    if not current_ticket:
                        code_match = re.search(r'tiket#\s*(\d+)', line.lower())
                        current_ticket = code_match.group(1) if code_match else None
            
            # Último reporte del archivo
            if current_ticket and current_report_text:
                hours = audit_text_minimal(current_report_text)
                reports[current_ticket] = max(reports.get(current_ticket, 0), hours)
    return reports

def audit_text_minimal(text):
    text = text.lower()
    # Redondeo para minutos en chat/fallback
    match_min = re.search(r'\((\d+)\s*min(?:utos)?\s*de\s*trabajo\)', text)
    if match_min: return math.ceil(int(match_min.group(1)) / 60)
    
    match_hours = re.search(r'\((\d+)\s*hor[as]+\s*de\s*trabajo\)', text)
    if match_hours: return int(match_hours.group(1))
    
    # Reutilizar reglas básicas
    if '58' in text and ('er' in text or 'error' in text): return 6
    if any(err in text for err in ['54', '55']) and ('er' in text or 'error' in text): return 5
    if '59' in text: return 1
    return 2

# --- LÓGICA 3: AUDITORÍA DE PDF ---
def parse_pdfs():
    reports = {} # {ticket_id: hours}
    pdf_files = glob.glob(os.path.join(PDF_DIR, "*.pdf"))
    
    for pdf_path in pdf_files:
        try:
            reader = PdfReader(pdf_path)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
            
            ticket_match = re.search(r'ticket\s*n[°º]?[:\s]+(\d+)', text, re.IGNORECASE)
            if ticket_match:
                ticket_id = ticket_match.group(1)
                hours = 0
                labor_match = re.search(r'tiempo de labor[:\s]+(\d+)[:\s]*(\d+)?\s*hs', text.lower())
                if labor_match:
                    hours = int(labor_match.group(1))
                else:
                    hours = audit_text_minimal(text)
                reports[ticket_id] = max(reports.get(ticket_id, 0), hours)
        except:
            continue
    return reports

# --- LÓGICA 4: AUDITORÍA DE BASE TÉCNICOS ---
def parse_base_tecnicos():
    reports = {} # {ticket_id: hours}
    if not os.path.exists(BASE_TECNICOS_PATH):
        print(f"Aviso: No se encontró {BASE_TECNICOS_PATH}")
        return reports
    
    try:
        df_base = pd.read_excel(BASE_TECNICOS_PATH)
        for _, row in df_base.iterrows():
            if pd.notna(row['Codigo']) and pd.notna(row['Horas de Trabajo']):
                # Asegurar que el ticket sea string limpio
                raw_codigo = str(row['Codigo'])
                if raw_codigo.endswith('.0'): raw_codigo = raw_codigo[:-2]
                reports[raw_codigo] = row['Horas de Trabajo']
    except Exception as e:
        print(f"Error parseando Base Tecnicos: {e}")
    
    return reports

# --- LÓGICA 5: AUDITORÍA DE MENSAJES DE RESOLUCIÓN ---
def get_resolution_hours(text):
    if pd.isna(text) or not isinstance(text, str):
        return 0
    
    text = text.lower()
    # Patrones para MINUTOS (Prioridad antes que horas)
    patterns_min = [
        r'horas?\s*realizadas?[:\s]*(\d+)\s*m\b',
        r'horas?\s*realizadas?[:\s]*(\d+)\s*min\b',
        r'(\d+)\s*min(?:utos)?\s*de\s*trabajo',
        r'duración\s*(?:del)?\s*trabajo\s*[:\s]*(\d+)\s*min\b',
        r'tiempo\s*de\s*labor\s*[:\s]*(\d+)\s*min\b',
        r'trabajo realizado en (\d+)\s*min\b'
    ]
    
    for pattern in patterns_min:
        match = re.search(pattern, text)
        if match:
            return math.ceil(int(match.group(1)) / 60)

    # Patrones para HORAS
    patterns_hs = [
        r'horas?\s*realizadas?[:\s]*(\d+)\s*h?',
        r'(\d+)\s*horas?\s*de\s*trabajo',
        r'duración\s*(?:del)?\s*trabajo\s*[:\s]*(\d+)\s*horas?',
        r'duración\s*[:\s]*(\d+)\s*horas?',
        r'tiempo\s*de\s*labor\s*[:\s]*(\d+)\s*horas?',
        r'(\d+)\s*hs\s*de\s*trabajo',
        r'\((\d+)horas? de trabajo\)',
        r'duración trabajo (\d+) horas?',
        r'trabajo realizado en (\d+)\s*h',
        r'(\d+)\s*h(?!\w)'
    ]
    
    for pattern in patterns_hs:
        match = re.search(pattern, text)
        if match:
            return int(match.group(1))
    
    return 0

# --- LÓGICA 6: PARSEO DE TARIFAS ---
def get_usd_rate():
    if not os.path.exists(TARIFAS_PATH):
        print(f"Aviso: No se encontró {TARIFAS_PATH}. Usando USD 25 por defecto.")
        return 25.0
    
    try:
        with open(TARIFAS_PATH, 'r', encoding='utf-8') as f:
            content = f.read().lower()
            # Buscar "valor 1 hora: usd 25" o similar
            match = re.search(r'valor 1 hora[:\s]*usd\s*(\d+)', content)
            if match:
                return float(match.group(1))
    except Exception as e:
        print(f"Error parseando tarifas: {e}")
    
    return 25.0

# --- MÉTODO PRINCIPAL ---
# --- MÉTODO PRINCIPAL ---
def main():
    print(f"--- Iniciando Auditoría Maestra ({MONTH_TO_EVAL.capitalize()}) ---")
    
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: No se encontró {EXCEL_PATH}")
        return

    # --- NUEVO: BACKUP AUTOMÁTICO ---
    try:
        if not os.path.exists(BACKUP_DIR):
            os.makedirs(BACKUP_DIR)
        
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = f"Target_Backup_{timestamp}.xlsx"
        backup_path = os.path.join(BACKUP_DIR, backup_filename)
        
        shutil.copy2(EXCEL_PATH, backup_path)
        print(f"Backup creado: {backup_path}")
    except Exception as e:
        print(f"Aviso: No se pudo realizar el backup: {e}")

    # Mapeo de meses en español
    meses = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8, 
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    target_month_num = meses.get(MONTH_TO_EVAL.lower(), 1)

    df_raw = pd.read_excel(EXCEL_PATH)
    
    # --- LIMPIEZA INICIAL: Evitar duplicación de TOTALES y registros vacíos ---
    # 1. Eliminar filas donde el Código sea nulo (filas fantasma)
    df_raw = df_raw.dropna(subset=['Código'])
    # 2. Eliminar filas que digan "TOTAL" (evita que el script procese sus propios totales anteriores)
    df_raw = df_raw[~df_raw['Código'].astype(str).str.contains('TOTAL', case=False, na=False)]
    
    # 1. Función para verificar Mes
    def is_in_month(date_val):
        if pd.isna(date_val): return False
        try:
            if isinstance(date_val, str):
                date_val = pd.to_datetime(date_val, dayfirst=True)
            return date_val.month == target_month_num
        except:
            return False

    df = df_raw.copy()
    print(f"Tickets a auditar: {len(df)}")

    out_of_month_tickets = []

    # Preparar columnas
    df['HS_Tecnica'] = 0
    df['HS_Chat'] = 0
    df['HS_PDF'] = 0
    df['base tecnicos'] = 0
    df['HS_Resolucion'] = 0 
    df['HS_Final'] = 0
    df['USD'] = 0.0
    df['Observaciones_Audit'] = ""
    df['Check'] = ""

    # Cargar datos de las fuentes (Skills)
    print("Ejecutando relevamiento de las 5 skills...")
    chat_results = parse_chats()
    pdf_results = parse_pdfs()
    base_results = parse_base_tecnicos()
    
    print(f"1. Skill Chat: {len(chat_results)} tickets encontrados.")
    print(f"2. Skill PDF: {len(pdf_results)} tickets encontrados.")
    print(f"3. Skill Base Técnicos: {len(base_results)} tickets encontrados.")
    print(f"4. Skill Mensaje de Resolución: Análisis directo del Excel.")
    print(f"5. Skill Reglas Técnicas: Aplicando a cada ticket.")
    
    usd_rate = get_usd_rate()
    print(f"Tarifa detectada: USD {usd_rate} por hora.")

    for idx, row in df.iterrows():
        ticket_id = str(row['Código'])
        
        # Verificar si está fuera de mes para la advertencia
        if not is_in_month(row['Fecha de Resolución']):
            fecha_str = str(row['Fecha de Resolución'])
            out_of_month_tickets.append(f"{ticket_id} ({fecha_str})")
        
        # Skill 1: Auditoría Técnica (Reglas de Diagnóstico)
        hs_tec = get_technical_hours(row)
        df.at[idx, 'HS_Tecnica'] = hs_tec
        
        # Skill 2: Auditoría de Chat
        hs_chat = chat_results.get(ticket_id, 0)
        df.at[idx, 'HS_Chat'] = hs_chat
        
        # Skill 3: Auditoría de PDF (Informes Formales)
        hs_pdf = pdf_results.get(ticket_id, 0)
        df.at[idx, 'HS_PDF'] = hs_pdf

        # Skill 4: Auditoría de Base de Técnicos
        hs_base = base_results.get(ticket_id, 0)
        df.at[idx, 'base tecnicos'] = hs_base

        # Skill 5: Auditoría de Mensaje de Resolución
        res_msg = row.get('Mensaje de Resolución', '')
        hs_res = get_resolution_hours(res_msg)
        df.at[idx, 'HS_Resolucion'] = hs_res
        
        # --- NUEVA REGLA: TICKET REPETIDO/DUPLICADO ---
        is_repeated = False
        full_context = (str(res_msg) + " " + str(row.get('Descripción', ''))).lower()
        if any(kw in full_context for kw in ['repetido', 'duplicado', 'triplicado', 'tkt duplicado', 'ticket duplicado']):
            is_repeated = True

        # Cálculo Final (Prioridad: Repetido > PDF > Base Técnicos > Mensaje Resolución > Chat > Técnica)
        if is_repeated:
            final_hours = 0
        elif hs_pdf > 0:
            final_hours = hs_pdf
        elif hs_base > 0:
            final_hours = hs_base
        elif hs_res > 0:
            final_hours = hs_res
        elif hs_chat > 0:
            final_hours = hs_chat
        else:
            final_hours = hs_tec
            
        df.at[idx, 'HS_Final'] = final_hours
        df.at[idx, 'USD'] = final_hours * usd_rate
        
        # Resumen de procedencia (Observaciones de Auditoría)
        obs = []
        if is_repeated:
            obs.append("Ticket Repetido/Duplicado")
        else:
            if hs_pdf > 0: obs.append("Skill: PDF")
            if hs_base > 0: obs.append("Skill: Base Técnicos")
            if hs_res > 0: obs.append("Skill: Resolución")
            if hs_chat > 0: obs.append("Skill: Chat")
            if hs_tec > 0: obs.append("Skill: Reglas Técnicas")
        
        if not obs:
            df.at[idx, 'Observaciones_Audit'] = "Sin datos en recursos"
        else:
            df.at[idx, 'Observaciones_Audit'] = " | ".join(obs)

    # --- LIMPIEZA Y REORGANIZACIÓN ---
    cols_to_drop = ['Tiempo Transcurrido (%)', 'Técnico Asignado', 'Desviación (%)', 'Título']
    df = df.drop(columns=[c for c in cols_to_drop if c in df.columns])
    
    # Ordenar Alfabéticamente por Sucursal
    sort_cols = [c for c in ['Sucursal', 'Cód. de Sucursal'] if c in df.columns]
    if sort_cols:
        df = df.sort_values(by=sort_cols).reset_index(drop=True)
    
    all_cols = list(df.columns)
    final_order = ['HS_Final', 'USD', 'Observaciones_Audit', 'Check']
    for col in final_order:
        if col in all_cols: all_cols.remove(col)
    for col in final_order:
        if col in df.columns: all_cols.append(col)
    df = df[all_cols]

    # --- GUARDAR CON ESTILOS (openpyxl) ---
    writer = pd.ExcelWriter(EXCEL_PATH, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Auditoria')
    
    workbook = writer.book
    worksheet = writer.sheets['Auditoria']
    
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.comments import Comment
    
    # 1. ESTILO DE ENCABEZADOS Y ALINEACIÓN GENERAL
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul oscuro
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Congelar paneles (Cabecera fija)
    worksheet.freeze_panes = "A2"
    
    # Aplicar Alineación Central y Bordes a TODAS las celdas con datos
    for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 2, min_col=1, max_col=len(all_cols)):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    # Aplicar Estilo de Cabecera (Sobrescribe alineación si es necesario, pero ya está centrada)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    # 2. IDENTIFICACIÓN DE COLUMNAS
    idx_codigo = all_cols.index('Código') + 1
    idx_hs = all_cols.index('HS_Final') + 1
    idx_usd = all_cols.index('USD') + 1
    idx_check = all_cols.index('Check') + 1
    
    col_hs = get_column_letter(idx_hs)
    col_usd = get_column_letter(idx_usd)
    col_check = get_column_letter(idx_check)
    
    total_row_idx = len(df) + 2
    last_data_idx = len(df) + 1
    
    # 3. Fila de TOTALES (Fórmulas)
    label_cell = worksheet.cell(row=total_row_idx, column=idx_codigo, value="TOTAL (Solo con X)")
    label_cell.font = Font(bold=True)
    
    range_check = f"${col_check}$2:${col_check}${last_data_idx}"
    
    hs_total_cell = worksheet.cell(row=total_row_idx, column=idx_hs, value=f'=SUMIF({range_check}, "X", {col_hs}2:{col_hs}{last_data_idx})')
    usd_total_cell = worksheet.cell(row=total_row_idx, column=idx_usd, value=f'=SUMIF({range_check}, "X", {col_usd}2:{col_usd}{last_data_idx})')
    
    # Estilo Fila Totales
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Azul muy suave
    for col_idx in range(1, len(all_cols) + 1):
        cell = worksheet.cell(row=total_row_idx, column=col_idx)
        cell.fill = total_fill
        cell.font = Font(bold=True)

    # 4. FORMATO DE MONEDA (USD)
    for row in range(2, total_row_idx + 1):
        worksheet.cell(row=row, column=idx_usd).number_format = '"$"#,##0.00'

    # 5. FORMATO CONDICIONAL (Fila Verde si Check == "X")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    
    # Aplicar a toda la fila de datos
    for row_idx in range(2, last_data_idx + 1):
        formula = f'${col_check}{row_idx}="X"'
        worksheet.conditional_formatting.add(
            f'A{row_idx}:{get_column_letter(len(all_cols))}{row_idx}',
            FormulaRule(formula=[formula], fill=green_fill, font=green_font)
        )

    # 6. AGREGAR COMENTARIOS (TOOLTIPS) PARA TEXTOS LARGOS
    cols_with_tooltips = ['Incidencia', 'Descripción', 'Última Respuesta', 'Mensaje de Resolución']
    tooltip_indices = [all_cols.index(c) + 1 for c in cols_with_tooltips if c in all_cols]
    
    for r_idx in range(2, last_data_idx + 1):
        for c_idx in tooltip_indices:
            cell = worksheet.cell(row=r_idx, column=c_idx)
            if cell.value and len(str(cell.value)) > 10:
                comment = Comment(str(cell.value), "Auditor Bot")
                comment.width = 450
                comment.height = 150
                cell.comment = comment

    # 7. AJUSTE DE COLUMNAS
    for idx, col_name in enumerate(all_cols, 1):
        letter = get_column_letter(idx)
        worksheet.column_dimensions[letter].width = 18

    writer.close()
    print(f"--- Proceso Finalizado con 5 Relevamientos ---")
    print(f"Total tickets analizados: {len(df)-1}")
    
    if out_of_month_tickets:
        print(f"\n[ADVERTENCIA] Se detectaron {len(out_of_month_tickets)} tickets fuera de {MONTH_TO_EVAL.capitalize()}:")
        # Mostrar solo los primeros 10 para no inundar la consola
        print(", ".join(out_of_month_tickets[:10]) + ("..." if len(out_of_month_tickets) > 10 else ""))
    
    print(f"\nArchivo actualizado: {EXCEL_PATH} (Base Completa)")

if __name__ == "__main__":
    main()
