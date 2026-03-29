import openpyxl
import re
import os
import glob
from difflib import get_close_matches
from datetime import datetime

# ─────────────────────────────────────────────────────────
# AUDITOR DE CHAT TÉCNICO — Esquema Completo
# Lee _chat.txt, matchea tickets con target.xlsx y escribe:
#   HS_Chat          → horas declaradas o estimadas del chat
#   HS_Final         → max(HS_Chat, Hs Mínimas del tarifario)
#   USD              → HS_Final × Valor Hora (USD)
#   Observaciones_Audit → fuente del cálculo
#   Check            → vacío, para revisión manual
# ─────────────────────────────────────────────────────────

SOURCE_DIR  = "/Users/CR1S714N/Documents/Repositorios GitHub/Franquicias/CALCULO DE HORAS/Source"
TARGET_PATH = f"{SOURCE_DIR}/target.xlsx"


# ─────────────────────────────────────────────────────────
# Extracción de horas desde texto de chat
# ─────────────────────────────────────────────────────────
def extraer_horas_chat(texto):
    """
    Extrae horas desde el texto del reporte. Reglas en orden de prioridad:
    1. Declarado: (X hora/s de trabajo)
    2. Tiempo de servicio: X hs / X hr / X min
    3. Reglas técnicas por código de error
    4. Default: 2 hs
    """
    t = texto.lower()
    fuente = "Reglas Técnicas"

    # 1. Horas declaradas explícitamente
    m = re.search(r'\((\d+(?:[.,]\d+)?)\s*hora[s]?\s*de\s*trabajo\)', t)
    if m:
        return float(m.group(1).replace(',', '.')), "Chat: Declarado"

    # 2. Tiempo de servicio
    m = re.search(r'tiempo de servicio[:\s]+(\d+(?:[.,]\d+)?)\s*(hs?|hr?|hora[s]?|min)?', t)
    if m:
        val = float(m.group(1).replace(',', '.'))
        unit = m.group(2) or 'hs'
        if unit.startswith('m'):
            val = max(1.0, round(val / 60, 1))
        return val, "Chat: Tiempo de Servicio"

    # 3. Reglas técnicas por error code
    if re.search(r'er\s*058|error\s*58', t):  return 6.0, "Reglas Técnicas: ER 058"
    if re.search(r'er\s*05[45]|error\s*5[45]', t): return 5.0, "Reglas Técnicas: ER 054/055"
    if re.search(r'er\s*059|error\s*59', t):  return 1.0, "Reglas Técnicas: ER 059"
    if 'puesta cero' in t or 'regulacion de muelas' in t: return 2.0, "Reglas Técnicas: Puesta Cero"

    return 2.0, "Default (sin dato explícito)"


# ─────────────────────────────────────────────────────────
# Parser del archivo de chat
# ─────────────────────────────────────────────────────────
def parse_chat(chat_path):
    """
    Parsea el _chat.txt de WhatsApp y devuelve lista de reportes técnicos.
    Cada reporte tiene: codigo, sucursal, fecha, texto, horas_declaradas
    """
    reports = []
    msg_start = re.compile(r'^\[(\d+/\d+/\d+),\s*[\d:]+\s*[ap]\.\s*m\.\]\s+(.+?):\s+(.*)')

    current = None

    with open(chat_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            match = msg_start.match(line)
            if match:
                fecha_str, sender, content = match.groups()
                # Ignora mensajes del sistema
                if any(x in content for x in ['imagen omitida', 'Se eliminó', 'cifrados de extremo', 'creó este grupo', 'añadió', 'añadieron']):
                    if current:
                        current['texto'] += ' ' + content
                    continue

                # Detectar inicio de reporte: tiene Tiket# o menciona una sucursal con horas
                es_reporte = bool(
                    re.search(r'tiket\s*#?\s*\d{5,}', content, re.IGNORECASE) or
                    re.search(r'\d+\s*hora[s]?\s*de\s*trabajo', content, re.IGNORECASE) or
                    re.search(r'cafetera\s+\w', content, re.IGNORECASE)
                )

                if es_reporte:
                    if current:
                        reports.append(current)
                    current = {
                        'fecha': fecha_str,
                        'texto': content,
                        'codigo': None,
                        'sucursal': None,
                    }
                elif current:
                    current['texto'] += ' ' + content
                    # Buscar código en continuación del mensaje
            elif current:
                current['texto'] += ' ' + line

    if current:
        reports.append(current)

    # Post-proceso: extraer código y sucursal de cada reporte
    for r in reports:
        t = r['texto']
        # Código de ticket
        code_m = re.search(r'tiket\s*#?\s*(\d{5,})', t, re.IGNORECASE)
        if code_m:
            r['codigo'] = code_m.group(1).strip()
        # Sucursal (primer línea tipo "Cafetera NOMBRE" o "NOMBRE (X horas)")
        suc_m = re.search(r'^(?:cafetera\s+)?([a-záéíóúüñ\s]+?)(?:\s*\(|\s+cimbali|\s+tiket|\s*$)', t, re.IGNORECASE)
        if suc_m:
            r['sucursal'] = suc_m.group(1).strip()

    return reports


# ─────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────
def main():
    # Buscar archivos de chat
    chat_files = sorted(glob.glob(os.path.join(SOURCE_DIR, "*chat*.txt")))
    if not chat_files:
        print("❌ No se encontraron archivos *chat*.txt en Source/")
        return
    print(f"📂 Archivos de chat: {[os.path.basename(f) for f in chat_files]}")

    # Parsear todos los chats
    all_reports = []
    for cf in chat_files:
        reps = parse_chat(cf)
        all_reports.extend(reps)
        print(f"   {os.path.basename(cf)}: {len(reps)} reportes extraídos")
    print(f"   Total: {len(all_reports)} reportes")

    # Cargar target
    wb = openpyxl.load_workbook(TARGET_PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    col_codigo = headers.index('Código')
    col_suc    = headers.index('Sucursal')

    # Obtener columnas de tarifario (ya integradas)
    col_vh  = headers.index('Valor Hora (USD)') if 'Valor Hora (USD)' in headers else None
    col_mnh = headers.index('Hs Mínimas')       if 'Hs Mínimas' in headers else None

    # Definir/crear columnas de auditoría
    AUDIT_COLS = ['HS_Chat', 'HS_Final', 'USD', 'Observaciones_Audit', 'Check']
    col_map = {}
    for col_name in AUDIT_COLS:
        if col_name in headers:
            col_map[col_name] = headers.index(col_name) + 1
        else:
            ws.cell(row=1, column=ws.max_column + 1, value=col_name)
            headers.append(col_name)
            col_map[col_name] = ws.max_column

    print(f"\n📊 Columnas de auditoría: {col_map}")

    # Leer todas las sucursales del target para fuzzy matching
    sucursales_target = {}  # nombre_lower → list of row_idx
    for row_idx in range(2, ws.max_row + 1):
        suc = ws.cell(row=row_idx, column=col_suc + 1).value
        if suc:
            key = str(suc).lower().strip()
            sucursales_target.setdefault(key, []).append(row_idx)

    # Procesar reportes
    mapeados = 0
    sin_match = 0

    for rep in all_reports:
        hs_chat, fuente = extraer_horas_chat(rep['texto'])
        rows_a_actualizar = []

        # 1. Match por código de ticket
        if rep['codigo']:
            for row_idx in range(2, ws.max_row + 1):
                cod_cell = ws.cell(row=row_idx, column=col_codigo + 1).value
                if cod_cell and str(cod_cell).strip() == str(rep['codigo']).strip():
                    rows_a_actualizar.append(row_idx)
                    break

        # 2. Si no hubo match, fuzzy por sucursal
        if not rows_a_actualizar and rep['sucursal']:
            suc_norm = rep['sucursal'].lower().strip()
            close = get_close_matches(suc_norm, list(sucursales_target.keys()), n=1, cutoff=0.45)
            if close:
                rows_a_actualizar = sucursales_target[close[0]]
                fuente += f" | Match sucursal: '{close[0]}'"

        if rows_a_actualizar:
            mapeados += 1
            for row_idx in rows_a_actualizar:
                # HS_Final = max(HS_Chat, Hs Mínimas del tarifario)
                hs_min = ws.cell(row=row_idx, column=col_mnh + 1).value if col_mnh is not None else 2
                vh     = ws.cell(row=row_idx, column=col_vh + 1).value  if col_vh  is not None else 35
                hs_min = hs_min or 2
                vh     = vh     or 35
                hs_final = max(hs_chat, float(hs_min))
                usd      = round(hs_final * float(vh), 2)

                ws.cell(row=row_idx, column=col_map['HS_Chat']).value          = hs_chat
                ws.cell(row=row_idx, column=col_map['HS_Final']).value         = hs_final
                ws.cell(row=row_idx, column=col_map['USD']).value              = usd
                ws.cell(row=row_idx, column=col_map['Observaciones_Audit']).value = fuente
                # Check vacío para revisión manual
        else:
            sin_match += 1

    wb.save(TARGET_PATH)

    print(f"\n✅ target.xlsx actualizado")
    print(f"   Mapeados:   {mapeados} de {len(all_reports)} reportes")
    print(f"   Sin match:  {sin_match} reportes")

    # Preview resultados
    print("\n📋 Muestra de filas actualizadas (con HS_Chat > 0):")
    wb2 = openpyxl.load_workbook(TARGET_PATH)
    ws2 = wb2.active
    h2  = [c.value for c in ws2[1]]
    shown = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        hs = row[h2.index('HS_Chat')] if 'HS_Chat' in h2 else None
        if hs and float(hs) > 0 and shown < 8:
            cod  = row[h2.index('Código')]
            suc  = row[h2.index('Sucursal')]
            fin  = row[h2.index('HS_Final')]
            usd  = row[h2.index('USD')]
            obs  = row[h2.index('Observaciones_Audit')]
            print(f"  [{cod}] {str(suc):<22} | HS_Chat={hs} | HS_Final={fin} | USD={usd} | {obs}")
            shown += 1


if __name__ == "__main__":
    main()

