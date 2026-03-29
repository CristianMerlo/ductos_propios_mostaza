import openpyxl
import os

# ──────────────────────────────────────────────
# TARIFARIO 2026 — Lógica de horas mínimas
# Tarifa técnica: 35 USD/hora (mano de obra pura)
#
# Reglas por distancia desde Congreso (km):
#   0  – 30 km  → 2 horas mínimas
#   30 – 60 km  → 3 horas mínimas
#   > 60 km     → 4 horas mínimas
# ──────────────────────────────────────────────

TARIFA_USD = 35  # Nueva tarifa 2026

def calcular_horas_minimas(km):
    """Devuelve las horas mínimas según distancia en km."""
    if km is None:
        return 4  # Default: máximo si no hay dato
    if km <= 30:
        return 2
    elif km <= 60:
        return 3
    else:
        return 4


def actualizar_tarifario():
    file_path = "/Users/CR1S714N/Documents/Repositorios GitHub/Franquicias/CALCULO DE HORAS/Tarifario/Clasificacion de Locales.xlsx"

    if not os.path.exists(file_path):
        print(f"Error: archivo no encontrado → {file_path}")
        return

    wb = openpyxl.load_workbook(file_path)
    hoja = wb["Hoja 1"]

    # Detectar índices de columnas en la fila de encabezado
    headers = {cell.value: cell.column for cell in hoja[1]}
    col_km   = headers.get("Kilometros")
    col_vh   = headers.get("VALOR HORA")
    col_minh = headers.get("MINIMO DE HORAS")

    if not col_km or not col_vh or not col_minh:
        print("Error: no se encontraron las columnas necesarias.")
        print("Columnas detectadas:", list(headers.keys()))
        return

    actualizados = 0
    for row in hoja.iter_rows(min_row=2):
        km_cell   = row[col_km - 1]
        vh_cell   = row[col_vh - 1]
        minh_cell = row[col_minh - 1]

        km = km_cell.value
        if km is None:
            continue  # Omitir filas sin dato de km

        horas_min = calcular_horas_minimas(km)

        vh_cell.value   = TARIFA_USD
        minh_cell.value = horas_min
        actualizados += 1

    wb.save(file_path)
    print(f"✅ Tarifario actualizado: {actualizados} locales procesados.")
    print(f"   Tarifa aplicada: {TARIFA_USD} USD/hora")
    print(f"   Archivo guardado: {file_path}")

    # Verificación rápida
    print("\n📋 Verificación (primeras 6 filas con datos):")
    wb2 = openpyxl.load_workbook(file_path)
    ws2 = wb2["Hoja 1"]
    count = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[col_km - 1] is not None and count < 6:
            local = row[0]
            km_v  = row[col_km - 1]
            vh_v  = row[col_vh - 1]
            min_v = row[col_minh - 1]
            print(f"   {local:<20} | {km_v:>7.1f} km | {vh_v} USD/hr | {min_v} hs mín")
            count += 1


def zona_desde_km(km):
    """Devuelve la etiqueta de zona según km."""
    if km is None:
        return "Zona 4 — Interior / Sin dato"
    if km <= 30:
        return "Zona 1 — CABA / GBA Cercano (≤30 km)"
    elif km <= 60:
        return "Zona 2 — GBA Medio (30-60 km)"
    else:
        return "Zona 3+ — GBA Lejano / Interior (>60 km)"


def generar_enriquecida():
    """
    Lee el Clasificacion de Locales.xlsx ya actualizado y genera
    el archivo Enriquecida con columnas de zona, tarifa y horas mínimas.
    """
    base_path    = "/Users/CR1S714N/Documents/Repositorios GitHub/Franquicias/CALCULO DE HORAS/Tarifario"
    input_file   = f"{base_path}/Clasificacion de Locales.xlsx"
    output_file  = f"{base_path}/Clasificacion de Locales Enriquecida.xlsx"

    if not os.path.exists(input_file):
        print(f"Error: archivo base no encontrado → {input_file}")
        return

    wb_in  = openpyxl.load_workbook(input_file)
    hoja   = wb_in["Hoja 1"]

    # Solo las primeras 6 columnas base: LOCAL, DIRECCION, LOCALIDAD, PROVINCIA, Kilometros, Radio
    COLS_BASE  = 6
    headers_in = [c.value for c in hoja[1]][:COLS_BASE]
    col_km_idx = headers_in.index("Kilometros")

    # Construir workbook de salida
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Tarifario Enriquecido"

    # Encabezados del archivo enriquecido
    nuevos_headers = headers_in + [
        "Radio/Zona Nueva",
        "VALOR HORA (USD)",
        "MINIMO DE HORAS",
        "Tarifa Técnica Mínima (USD)"
    ]
    ws_out.append(nuevos_headers)

    filas_procesadas = 0
    for row in hoja.iter_rows(min_row=2, values_only=True):
        base_row = list(row[:COLS_BASE])
        km = row[col_km_idx]
        if km is None:
            ws_out.append(base_row + ["Sin dato", TARIFA_USD, 4, TARIFA_USD * 4])
            continue

        horas      = calcular_horas_minimas(km)
        zona       = zona_desde_km(km)
        tarifa_min = TARIFA_USD * horas

        ws_out.append(base_row + [zona, TARIFA_USD, horas, tarifa_min])
        filas_procesadas += 1

    wb_out.save(output_file)
    print(f"✅ Enriquecida generada: {filas_procesadas} locales procesados.")
    print(f"   Archivo: {output_file}")

    # Preview
    print("\n📋 Preview (5 filas):")
    col_zona = nuevos_headers.index("Radio/Zona Nueva")
    col_vh   = nuevos_headers.index("VALOR HORA (USD)")
    col_mh   = nuevos_headers.index("MINIMO DE HORAS")
    col_tm   = nuevos_headers.index("Tarifa Técnica Mínima (USD)")
    wb_check = openpyxl.load_workbook(output_file)
    ws_check = wb_check.active
    count = 0
    for row in ws_check.iter_rows(min_row=2, values_only=True):
        if count >= 5:
            break
        print(f"   {str(row[0]):<20} | {str(row[col_km_idx]):>7} km | {row[col_zona]:<45} | {row[col_vh]} USD | {row[col_mh]} hs mín | {row[col_tm]} USD mín")
        count += 1


if __name__ == "__main__":
    actualizar_tarifario()
    print()
    generar_enriquecida()
