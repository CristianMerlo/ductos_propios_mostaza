import pandas as pd
import math
import shutil
import re
import os
import glob
import json
from datetime import datetime
from pypdf import PdfReader

# ─────────────────────────────────────────────────────────
# CONFIGURACIÓN — paths relativos al script
# ─────────────────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
SOURCE_DIR = os.path.join(BASE_DIR, "Source")
PDF_DIR    = os.path.join(SOURCE_DIR, "Informes")
BACKUP_DIR = os.path.join(SOURCE_DIR, "Backup")
BASE_TECNICOS_PATH = os.path.join(SOURCE_DIR, "Base Tecnicos.xlsx")
PARAMETROS_PATH    = os.path.join(SOURCE_DIR, "PARAMETROS.txt")
SCHEMA_PATH        = os.path.join(SOURCE_DIR, "SCHEMA.json")


def load_schema():
    """Carga el mapeo de columnas desde SCHEMA.json."""
    if os.path.exists(SCHEMA_PATH):
        with open(SCHEMA_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    print("⚠️  No se encontró SCHEMA.json. Se usarán nombres por defecto.")
    return {}


def find_target_excel():
    """Encuentra el archivo target sin importar el case del nombre."""
    if not os.path.exists(SOURCE_DIR):
        return None
    for fname in os.listdir(SOURCE_DIR):
        if fname.lower().replace(" ", "") == "target.xlsx":
            return os.path.join(SOURCE_DIR, fname)
    # Buscar cualquier variante
    for fname in os.listdir(SOURCE_DIR):
        if "target" in fname.lower() and fname.endswith(".xlsx") and "~$" not in fname and "old" not in fname.lower():
            return os.path.join(SOURCE_DIR, fname)
    return None


# ─────────────────────────────────────────────────────────
# SKILL 1 — REGLAS TÉCNICAS
# ─────────────────────────────────────────────────────────
def get_technical_hours(row, schema):
    """Calcula horas por reglas técnicas desde el contenido del ticket."""
    # Usar nombres del esquema
    fields = [
        schema.get('titulo', 'Título'),
        schema.get('incidencia', 'Incidencia'),
        schema.get('descripcion', 'Descripción'),
        schema.get('ultima_respuesta', 'Última Respuesta'),
        schema.get('mensaje_resolucion', 'Mensaje de Resolución')
    ]
    text = " ".join(str(row.get(f, "") or "") for f in fields if f in row.index).lower()

    # Resuelto por el local → 0 hs
    if any(p in text for p in ['resuelto por el local', 'solucionado por el local',
                                'local soluciono', 'local lo soluciono', 'solucionado por local']):
        return 0

    # Errores Cimbali
    if any(e in text for e in ['54', '55']) and any(w in text for w in ['er', 'error']): return 5
    if '58' in text and any(w in text for w in ['er', 'error']): return 6
    if '59' in text: return 1

    # Equipos no-café
    if any(w in text for w in ['freidora', 'locatelli']):
        if any(w in text for w in ['picos', 'quemadores', 'explos']): return 3
    if any(w in text for w in ['horno', 'unox']):
        if any(w in text for w in ['contactora', 'cable']): return 5
    if any(w in text for w in ['broiler', 'nieco']): return 2

    # Triviales
    if any(w in text for w in ['destrabe', 'madera']): return 3
    if 'se cierra' in text and len(text.strip()) < 50: return 2

    # Tiempo de servicio explícito
    m = re.search(r'tiempo de servicio[:\s]+(\d+)\s*([a-z]+)?', text)
    if m:
        val = int(m.group(1))
        unit = m.group(2) or 'hs'
        return math.ceil(val / 60) if unit.startswith('m') else val

    return 2 if len(text.strip()) > 10 else 0


# ─────────────────────────────────────────────────────────
# SKILL 2 — AUDITORÍA DE CHAT
# ─────────────────────────────────────────────────────────
def audit_text_minimal(text):
    t = text.lower()
    m = re.search(r'\((\d+)\s*min(?:utos)?\s*de\s*trabajo\)', t)
    if m: return math.ceil(int(m.group(1)) / 60)
    m = re.search(r'\((\d+(?:[.,]\d+)?)\s*hora[s]?\s*de\s*trabajo\)', t)
    if m: return float(m.group(1).replace(',', '.'))
    if '58' in t and ('er' in t or 'error' in t): return 6
    if any(e in t for e in ['54', '55']) and ('er' in t or 'error' in t): return 5
    if '59' in t: return 1
    return 2


def parse_chats():
    """Lee todos los *chat*.txt de Source/ y devuelve {ticket_id: hours}."""
    reports = {}
    chat_files = glob.glob(os.path.join(SOURCE_DIR, "*chat*.txt"))
    msg_start = re.compile(r'^\[(\d+/\d+/\d+),.*\] (.*): (.*)')

    for chat_path in chat_files:
        current_text = ""
        current_ticket = None
        with open(chat_path, 'r', encoding='utf-8') as f:
            for line in f:
                m = msg_start.match(line)
                if m:
                    if current_ticket and current_text:
                        hours = audit_text_minimal(current_text)
                        reports[current_ticket] = max(reports.get(current_ticket, 0), hours)
                    _, _, content = m.groups()
                    current_text = content
                    cm = re.search(r'tiket\s*#?\s*(\d+)', content, re.IGNORECASE)
                    current_ticket = cm.group(1) if cm else None
                else:
                    current_text += " " + line.strip()
                    if not current_ticket:
                        cm = re.search(r'tiket\s*#?\s*(\d+)', line, re.IGNORECASE)
                        current_ticket = cm.group(1) if cm else None
            if current_ticket and current_text:
                hours = audit_text_minimal(current_text)
                reports[current_ticket] = max(reports.get(current_ticket, 0), hours)
    return reports


# ─────────────────────────────────────────────────────────
# SKILL 3 — ANÁLISIS DE PDFs
# ─────────────────────────────────────────────────────────
def parse_pdfs():
    """Lee PDFs de Source/Informes/ y devuelve {ticket_id: hours}."""
    reports = {}
    if not os.path.exists(PDF_DIR):
        return reports
    for pdf_path in glob.glob(os.path.join(PDF_DIR, "*.pdf")):
        try:
            reader = PdfReader(pdf_path)
            text = "\n".join(p.extract_text() or "" for p in reader.pages)
            m_ticket = re.search(r'ticket\s*n[°º]?[:\s]+(\d+)', text, re.IGNORECASE)
            if m_ticket:
                ticket_id = m_ticket.group(1)
                hours = 0
                m_labor = re.search(r'tiempo de labor[:\s]+(\d+)[:\s]*(\d+)?\s*hs', text.lower())
                if m_labor:
                    hours = int(m_labor.group(1))
                else:
                    hours = audit_text_minimal(text)
                reports[ticket_id] = max(reports.get(ticket_id, 0), hours)
        except Exception:
            continue
    return reports


# ─────────────────────────────────────────────────────────
# SKILL 4 — BASE DE TÉCNICOS
# ─────────────────────────────────────────────────────────
def parse_base_tecnicos():
    """Lee Base Tecnicos.xlsx y devuelve {ticket_id: hours}."""
    reports = {}
    if not os.path.exists(BASE_TECNICOS_PATH):
        return reports
    try:
        df_base = pd.read_excel(BASE_TECNICOS_PATH)
        # Buscar columna de código/horas dinámicamente o usar default
        col_cod = next((c for c in df_base.columns if "cod" in c.lower()), df_base.columns[0])
        col_hs  = next((c for c in df_base.columns if "hora" in c.lower()), df_base.columns[-1])
        
        for _, row in df_base.iterrows():
            if pd.notna(row.get(col_cod)) and pd.notna(row.get(col_hs)):
                raw = str(row[col_cod])
                if raw.endswith('.0'): raw = raw[:-2]
                reports[raw] = row[col_hs]
    except Exception as e:
        print(f"  ⚠️  Error Base Técnicos: {e}")
    return reports


# ─────────────────────────────────────────────────────────
# SKILL 5 — MENSAJE DE RESOLUCIÓN
# ─────────────────────────────────────────────────────────
def get_resolution_hours(text):
    """Extrae horas declaradas en el Mensaje de Resolución."""
    if not text or not isinstance(text, str): return 0
    t = text.lower()

    patterns_min = [
        r'horas?\s*realizadas?[:\s]*(\d+)\s*m\b',
        r'(\d+)\s*min(?:utos)?\s*de\s*trabajo',
        r'tiempo\s*de\s*labor\s*[:\s]*(\d+)\s*min\b',
    ]
    for p in patterns_min:
        m = re.search(p, t)
        if m: return math.ceil(int(m.group(1)) / 60)

    patterns_hs = [
        r'horas?\s*realizadas?[:\s]*(\d+)',
        r'(\d+)\s*horas?\s*de\s*trabajo',
        r'duración\s*(?:del)?\s*trabajo\s*[:\s]*(\d+)\s*horas?',
        r'tiempo\s*de\s*labor\s*[:\s]*(\d+)\s*horas?',
        r'(\d+)\s*hs\s*de\s*trabajo',
        r'\((\d+)horas? de trabajo\)',
        r'tiempo de servicio[:\s]+(\d+)',
    ]
    for p in patterns_hs:
        m = re.search(p, t)
        if m: return int(m.group(1))
    return 0


# ─────────────────────────────────────────────────────────
# MÉTODO PRINCIPAL
# ─────────────────────────────────────────────────────────
def main():
    print("─" * 60)
    print("  AUDITOR MAESTRO — 5 Skills (Esquema Dinámico)")
    print("─" * 60)

    # Cargar esquema
    schema = load_schema()
    col_codigo = schema.get('codigo', 'Código')
    col_resolucion = schema.get('mensaje_resolucion', 'Mensaje de Resolución')
    col_sucursal = schema.get('sucursal', 'Sucursal')
    col_vh = schema.get('valor_hora', 'Valor Hora (USD)')
    col_min = schema.get('hs_minimas', 'Hs Mínimas')

    # Encontrar target
    excel_path = find_target_excel()
    if not excel_path:
        print(f"❌ No se encontró ningún target.xlsx en {SOURCE_DIR}")
        return
    print(f"✅ Target: {os.path.basename(excel_path)}")

    # Backup automático
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_path = os.path.join(BACKUP_DIR, f"Target_Backup_{ts}.xlsx")
    shutil.copy2(excel_path, backup_path)
    print(f"💾 Backup: {os.path.basename(backup_path)}")

    # Cargar datos
    df_raw = pd.read_excel(excel_path)
    if col_codigo not in df_raw.columns:
        print(f"❌ Error: La columna '{col_codigo}' no existe en el Excel.")
        return
        
    df_raw = df_raw.dropna(subset=[col_codigo])
    df_raw = df_raw[~df_raw[col_codigo].astype(str).str.contains('TOTAL', case=False, na=False)]
    df = df_raw.copy()
    print(f"📊 Tickets a auditar: {len(df)}")

    # Ejecutar fuentes
    print("\n🔍 Ejecutando skills...")
    chat_results  = parse_chats()
    pdf_results   = parse_pdfs()
    base_results  = parse_base_tecnicos()

    audit_cols = {
        'HS_Tecnica':    schema.get('hs_tecnica', 'HS_Tecnica'),
        'HS_Chat':       schema.get('hs_chat', 'HS_Chat'),
        'HS_PDF':        schema.get('hs_pdf', 'HS_PDF'),
        'base tecnicos': schema.get('base_tecnicos', 'base tecnicos'),
        'HS_Resolucion': schema.get('hs_resolucion', 'HS_Resolucion'),
        'HS_Final':      schema.get('hs_final', 'HS_Final'),
        'USD':           schema.get('usd', 'USD'),
        'Observaciones_Audit': schema.get('observaciones_audit', 'Observaciones_Audit'),
        'Check':         schema.get('check', 'Check')
    }

    # Inicializar columnas
    for role, col_name in audit_cols.items():
        if col_name not in df.columns:
            df[col_name] = 0 if role not in ['Observaciones_Audit', 'Check'] else ""

    # Procesar fila por fila
    for idx, row in df.iterrows():
        ticket_id = str(int(row[col_codigo])) if not pd.isna(row[col_codigo]) else ""

        hs_tec  = get_technical_hours(row, schema)
        hs_chat = chat_results.get(ticket_id, 0)
        hs_pdf  = pdf_results.get(ticket_id, 0)
        hs_base = base_results.get(ticket_id, 0)
        hs_res  = get_resolution_hours(row.get(col_resolucion, ''))

        df.at[idx, audit_cols['HS_Tecnica']]    = hs_tec
        df.at[idx, audit_cols['HS_Chat']]       = hs_chat
        df.at[idx, audit_cols['HS_PDF']]        = hs_pdf
        df.at[idx, audit_cols['base tecnicos']] = hs_base
        df.at[idx, audit_cols['HS_Resolucion']] = hs_res

        # Detectar repetidos/duplicados
        desc_text = str(row.get(schema.get('descripcion', 'Descripción'), ''))
        res_text  = str(row.get(col_resolucion, ''))
        full_context = (res_text + " " + desc_text).lower()
        is_repeated = any(kw in full_context for kw in
                          ['repetido', 'duplicado', 'triplicado', 'tkt duplicado', 'ticket duplicado'])

        # Resuelto por local
        fields_for_local = [
            schema.get('incidencia', 'Incidencia'),
            schema.get('descripcion', 'Descripción'),
            schema.get('ultima_respuesta', 'Última Respuesta'),
            col_resolucion
        ]
        text_local = " ".join(str(row.get(f, "") or "") for f in fields_for_local if f in row.index).lower()
        is_local   = any(p in text_local for p in
                         ['resuelto por el local', 'solucionado por el local',
                          'local soluciono', 'local lo soluciono', 'solucionado por local'])

        if is_repeated:
            final_hours, obs_tags = 0, ["Ticket Repetido/Duplicado"]
        elif is_local:
            final_hours, obs_tags = 0, ["Resuelto por el Local (0 HS)"]
        else:
            obs_tags = []
            if hs_pdf  > 0: obs_tags.append("Skill: PDF")
            if hs_base > 0: obs_tags.append("Skill: Base Técnicos")
            if hs_res  > 0: obs_tags.append("Skill: Resolución")
            if hs_chat > 0: obs_tags.append("Skill: Chat")
            if hs_tec  > 0: obs_tags.append("Skill: Reglas Técnicas")

            final_hours = (hs_pdf  if hs_pdf  > 0 else
                           hs_base if hs_base > 0 else
                           hs_res  if hs_res  > 0 else
                           hs_chat if hs_chat > 0 else hs_tec)

        # Aplicar mínimo de horas
        hs_min = row.get(col_min, 0)
        if final_hours > 0 and hs_min and not pd.isna(hs_min):
            final_hours = max(final_hours, float(hs_min))

        df.at[idx, audit_cols['HS_Final']] = final_hours
        df.at[idx, audit_cols['Observaciones_Audit']] = " | ".join(obs_tags) if obs_tags else "Sin datos"

    # Tarifa
    usd_rate = 35.0
    if col_vh in df.columns:
        sample = df[col_vh].dropna()
        if not sample.empty:
            usd_rate = float(sample.iloc[0])
    print(f"\n💵 Tarifa detectada: USD {usd_rate}/hora")

    # Guardar con estilos
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import FormulaRule

    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Auditoria')
    wb, ws = writer.book, writer.sheets['Auditoria']

    all_cols, last_row = list(df.columns), len(df) + 1

    # Fórmulas USD
    if audit_cols['HS_Final'] in all_cols and audit_cols['USD'] in all_cols:
        col_idx_hs = all_cols.index(audit_cols['HS_Final']) + 1
        col_idx_usd = all_cols.index(audit_cols['USD']) + 1
        letter_hs = get_column_letter(col_idx_hs)
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=col_idx_usd).value = f'={letter_hs}{r}*{usd_rate}'
            ws.cell(row=r, column=col_idx_usd).number_format = '"$"#,##0.00'

    # Estética
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=1, max_row=last_row + 1, min_col=1, max_col=len(all_cols)):
        for cell in row:
            cell.alignment = center
            cell.border = thin_border

    for cell in ws[1]:
        cell.fill, cell.font = header_fill, header_font

    # Tooltips
    from openpyxl.worksheet.datavalidation import DataValidation
    TOOLTIP_ROLES = ['descripcion', 'ultima_respuesta', 'mensaje_resolucion', 'observaciones_audit']
    for role in TOOLTIP_ROLES:
        c_name = schema.get(role)
        if c_name and c_name in all_cols:
            c_idx = all_cols.index(c_name) + 1
            for r in range(2, last_row + 1):
                cell = ws.cell(row=r, column=c_idx)
                if cell.value and len(str(cell.value)) > 10:
                    dv = DataValidation(type="custom", formula1="=TRUE", allow_blank=True, showInputMessage=True)
                    dv.promptTitle, dv.prompt = f"📋 {c_name}", str(cell.value)[:250]
                    ws.add_data_validation(dv)
                    dv.add(cell)

    writer.close()
    print(f"✅ Proceso finalizado. Archivo guardado: {os.path.basename(excel_path)}")


if __name__ == "__main__":
    main()
